from django.http import JsonResponse
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from datetime import datetime
from Area_comercial.models import *
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook
from django.http import HttpResponse
from collections import defaultdict, Counter
from decimal import Decimal
from django.http import HttpResponse
import openpyxl
from Area_comercial.api.serializer import *


class Exportecomercial(APIView):
    def get(self, request):
        empresa_data = Empresa.objects.all()
        ofertas_data = Ofertas.objects.all()
        notas_data = Notas.objects.all()
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exporte comercial"
        headers = ["Empresa"]
        
        ws.append(headers)

        for oferta in ofertas_data:
                # ws.append([empresa.Nombre_empresa])
                ws.append([oferta.Nuevo,oferta.Descripcion,oferta.Estado,oferta.Pago_mensual,oferta.Sector,oferta.Causal_negacion,oferta.Canal_medio])
                
      

        temp_file = 'Area_comercial_temp.xlsx'
        wb.save(temp_file)

       
        wb2 = load_workbook(temp_file)
       
        wb2.save(temp_file)

        
        with open(temp_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Exporte_comercial.xlsx"'

        return response


        
        
        
 